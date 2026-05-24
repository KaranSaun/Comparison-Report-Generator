#!/usr/bin/env python3
"""
Warehouse Data Reconciliation Engine
=====================================
Compares WMS (Warehouse Management System) and WCS (Warehouse Control System)
Excel exports, matches records by a concatenated ID, compares quantities, and
generates a formatted Comparison_Report.xlsx with multiple audit sheets.

Tech stack: Python · pandas · openpyxl
"""

import logging
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Logging setup — console + file
# ---------------------------------------------------------------------------
LOG_FILE = "reconciliation.log"

logger = logging.getLogger("reconciliation")
logger.setLevel(logging.DEBUG)

_console_handler = logging.StreamHandler()
_console_handler.setLevel(logging.INFO)
_console_handler.setFormatter(
    logging.Formatter("%(asctime)s | %(levelname)-8s | %(message)s", "%Y-%m-%d %H:%M:%S")
)

_file_handler = logging.FileHandler(LOG_FILE, mode="a", encoding="utf-8")
_file_handler.setLevel(logging.DEBUG)
_file_handler.setFormatter(
    logging.Formatter("%(asctime)s | %(levelname)-8s | %(message)s", "%Y-%m-%d %H:%M:%S")
)

logger.addHandler(_console_handler)
logger.addHandler(_file_handler)

# ---------------------------------------------------------------------------
# Configuration (easy to modify for future changes)
# ---------------------------------------------------------------------------
WCS_COLUMNS = {
    "barcode": "Carton Barcode",
    "item": "Item",
    "qty": "PutQty",
}
WMS_COLUMNS = {
    "id": "ID",
    "sku": "SKU",
    "qty": "Qty",
}
WCS_REQUIRED_COLS = set(WCS_COLUMNS.values())
WMS_REQUIRED_COLS = set(WMS_COLUMNS.values())

# ---------------------------------------------------------------------------
# Formatting constants
# ---------------------------------------------------------------------------
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ===========================  LOADING  =====================================

def _detect_header_row(filepath: str, required_cols: set, max_scan: int = 20) -> int:
    """
    Scan the first *max_scan* rows of an Excel file and return the 0-indexed
    row number that contains all *required_cols*.  This handles WMS files that
    may have metadata rows above the real header.

    Uses case-insensitive + trimmed comparison for failsafe column matching.
    """
    preview = pd.read_excel(
        filepath,
        header=None,
        nrows=max_scan,
        dtype=str,
        keep_default_na=False,
        engine="openpyxl",
    )

    # Normalize required cols for case-insensitive matching
    required_lower = {c.strip().lower() for c in required_cols}

    for idx, row in preview.iterrows():
        row_values = {str(v).strip().lower() for v in row.values if v != ""}
        if required_lower.issubset(row_values):
            logger.info("Detected header row at index %d in %s", idx, filepath)
            return int(idx)

    raise ValueError(
        f"Could not find required columns {required_cols} in the first "
        f"{max_scan} rows of '{filepath}'. Please check the file."
    )


def load_excel(
    filepath: str,
    required_cols: set,
    header_row: Optional[int] = None,
) -> pd.DataFrame:
    """
    Load an Excel file, auto-detecting the header row if not specified.
    All cells are read as strings to prevent dtype corruption.
    keep_default_na=False prevents pandas from converting barcode strings to NaN.
    """
    filepath = str(filepath)
    logger.info("Loading file: %s", filepath)

    if not Path(filepath).is_file():
        raise FileNotFoundError(f"File not found: {filepath}")

    # Auto-detect header if not provided
    if header_row is None:
        header_row = _detect_header_row(filepath, required_cols)

    df = pd.read_excel(
        filepath,
        header=header_row,
        dtype=str,               # CRITICAL: prevents float/scientific notation
        keep_default_na=False,   # CRITICAL: preserves barcode strings like "NA..."
        engine="openpyxl",
    )

    # Failsafe header cleanup: strip whitespace from column names
    df.columns = df.columns.str.strip()

    # Validate that all required columns exist (case-insensitive)
    validate_columns(df, required_cols, filepath)

    logger.info(
        "Loaded %d rows × %d cols from %s (header at row %d)",
        len(df), len(df.columns), filepath, header_row,
    )
    return df


# ===========================  VALIDATION  ==================================

def validate_columns(df: pd.DataFrame, required: set, source: str) -> None:
    """
    Raise a readable error if any required columns are missing.
    Uses case-insensitive + trimmed matching for failsafe detection.
    """
    # Build a map: lowercase_trimmed → actual column name
    col_map = {c.strip().lower(): c for c in df.columns}
    required_lower = {c.strip().lower() for c in required}

    missing = required_lower - set(col_map.keys())
    if missing:
        # Report in original casing for clarity
        raise KeyError(
            f"Missing columns in '{source}': {missing}. "
            f"Available columns: {list(df.columns)}"
        )

    # Rename columns to the canonical required names (preserves expected casing)
    rename_map = {}
    for req in required:
        req_lower = req.strip().lower()
        actual = col_map.get(req_lower)
        if actual and actual != req:
            rename_map[actual] = req
    if rename_map:
        df.rename(columns=rename_map, inplace=True)
        logger.debug("Renamed columns for consistency: %s", rename_map)


def validate_no_null_keys(df: pd.DataFrame, key_col: str, source: str) -> None:
    """Warn about rows where the matching key is null after cleaning."""
    null_mask = df[key_col].isna() | (df[key_col].astype(str).str.strip() == "")
    n_null = null_mask.sum()
    if n_null:
        logger.warning(
            "%d row(s) in %s have null/empty '%s' — these will NOT match.",
            n_null, source, key_col,
        )


def validate_id_format(df: pd.DataFrame, key_col: str, source: str) -> None:
    """
    Detect malformed IDs — those that have empty segments before or after the
    hyphen (e.g. '-800492-000141' or 'B0002655835-').
    """
    ids = df[key_col].astype(str)
    malformed_mask = (
        ids.str.startswith("-") |
        ids.str.endswith("-") |
        ids.str.contains("--", regex=False)
    )
    n_bad = malformed_mask.sum()
    if n_bad:
        samples = ids[malformed_mask].head(5).tolist()
        logger.warning(
            "%d malformed ID(s) in %s (samples: %s). "
            "These may fail to match correctly.",
            n_bad, source, samples,
        )


def validate_duplicates(
    df: pd.DataFrame, key_col: str, source: str
) -> pd.DataFrame:
    """
    Detect and log duplicate IDs. Returns a DataFrame of duplicate details
    for the audit sheet.
    """
    dup_counts = df[key_col].value_counts()
    dup_counts = dup_counts[dup_counts > 1]
    if len(dup_counts):
        logger.warning(
            "%d unique duplicate '%s' entries found in %s (%d total rows). "
            "Quantities will be summed per ID.",
            len(dup_counts), key_col, source, dup_counts.sum(),
        )
    return df


def get_duplicate_details(
    df: pd.DataFrame, key_col: str, source_name: str
) -> pd.DataFrame:
    """Build a DataFrame of duplicate IDs with counts for the Duplicate_IDs sheet."""
    dup_counts = df[key_col].value_counts()
    dup_counts = dup_counts[dup_counts > 1].reset_index()
    dup_counts.columns = ["Duplicate ID", "Count"]
    dup_counts["Source"] = source_name
    return dup_counts[["Source", "Duplicate ID", "Count"]]


def validate_row_counts(wcs_df: pd.DataFrame, wms_df: pd.DataFrame) -> None:
    """Log row-count comparison between the two sources."""
    logger.info(
        "Row counts — WCS: %d | WMS: %d | Δ: %d",
        len(wcs_df), len(wms_df), abs(len(wcs_df) - len(wms_df)),
    )


# ===========================  CLEANING  ====================================

def normalize_string(s: pd.Series) -> pd.Series:
    """
    Advanced ID normalization layer:
      - Force str dtype
      - Unicode NFC normalization
      - Remove invisible/control characters
      - Collapse multiple spaces into single space
      - Replace multiple hyphens with single hyphen
      - Strip leading/trailing whitespace
      - Remove non-breaking spaces
      - Preserve leading zeros (string throughout)
    """
    s = s.astype(str)
    # Unicode normalization (NFC form)
    s = s.apply(lambda x: unicodedata.normalize("NFC", x))
    # Remove invisible/control characters (categories C*)
    s = s.apply(
        lambda x: "".join(ch for ch in x if not unicodedata.category(ch).startswith("C") or ch in ("\n", "\r", "\t"))
    )
    # Remove newlines, tabs, non-breaking spaces
    s = s.str.replace(r"[\r\n\t]", "", regex=True)
    s = s.str.replace("\u00a0", " ", regex=False)
    # Collapse multiple spaces into one
    s = s.str.replace(r"\s+", " ", regex=True)
    # Replace multiple consecutive hyphens with single hyphen
    s = s.str.replace(r"-{2,}", "-", regex=True)
    # Final strip
    s = s.str.strip()
    # Handle empty/nan literals
    s = s.replace({"nan": "", "None": "", "<NA>": "", "NaN": ""})
    return s


def clean_qty_column(series: pd.Series) -> pd.Series:
    """
    Safely convert a quantity column to integer.
    Non-numeric values become NaN → filled with 0.
    Result is int to avoid 1.0/2.0 in Excel output.
    """
    numeric = pd.to_numeric(series.astype(str).str.strip(), errors="coerce").fillna(0)
    return numeric.astype(int)


def clean_wcs(df: pd.DataFrame) -> pd.DataFrame:
    """Clean WCS-specific columns using vectorized operations."""
    df = df.copy()
    df[WCS_COLUMNS["barcode"]] = normalize_string(df[WCS_COLUMNS["barcode"]])
    df[WCS_COLUMNS["item"]] = normalize_string(df[WCS_COLUMNS["item"]])
    df[WCS_COLUMNS["qty"]] = clean_qty_column(df[WCS_COLUMNS["qty"]])
    return df


def clean_wms(df: pd.DataFrame) -> pd.DataFrame:
    """Clean WMS-specific columns using vectorized operations."""
    df = df.copy()
    df[WMS_COLUMNS["id"]] = normalize_string(df[WMS_COLUMNS["id"]])
    df[WMS_COLUMNS["sku"]] = normalize_string(df[WMS_COLUMNS["sku"]])
    df[WMS_COLUMNS["qty"]] = clean_qty_column(df[WMS_COLUMNS["qty"]])
    return df


# ===========================  CONCATENATION  ===============================

def build_concat_id(col_a: pd.Series, col_b: pd.Series) -> pd.Series:
    """
    Concatenate two string columns with a single hyphen separator.
    Format: <col_a>-<col_b>
    - No extra spaces
    - No dtype corruption
    - Case-insensitive (uppercased)
    """
    return col_a.str.strip().str.upper() + "-" + col_b.str.strip().str.upper()


def add_wcs_id(df: pd.DataFrame) -> pd.DataFrame:
    """Create the common ID on the WCS dataframe: <Carton Barcode>-<Item>"""
    df = df.copy()
    df["ID"] = build_concat_id(df[WCS_COLUMNS["barcode"]], df[WCS_COLUMNS["item"]])
    return df


def add_wms_id(df: pd.DataFrame) -> pd.DataFrame:
    """Create the common ID on the WMS dataframe: <ID>-<SKU>"""
    df = df.copy()
    df["ID"] = build_concat_id(df[WMS_COLUMNS["id"]], df[WMS_COLUMNS["sku"]])
    return df


# ===========================  MATCHING  ====================================

def aggregate_quantities(
    df: pd.DataFrame, id_col: str, qty_col: str
) -> pd.DataFrame:
    """Sum quantities per ID to handle duplicate rows safely (vectorized)."""
    return (
        df.groupby(id_col, as_index=False)[qty_col]
        .sum()
        .rename(columns={qty_col: qty_col})
    )


def match_datasets(
    wcs_df: pd.DataFrame, wms_df: pd.DataFrame
) -> pd.DataFrame:
    """
    Full OUTER join on the concatenated ID (VLOOKUP/XLOOKUP equivalent).
    Returns a dataframe with columns: ID, WCS Qty, WMS Qty, Status.

    Status hierarchy:
      - Missing in WMS  (ID only in WCS)
      - Missing in WCS  (ID only in WMS)
      - Matched         (quantities equal)
      - Unmatched       (quantities differ)
    """
    # Aggregate in case of duplicates (vectorized)
    wcs_agg = aggregate_quantities(wcs_df, "ID", WCS_COLUMNS["qty"])
    wms_agg = aggregate_quantities(wms_df, "ID", WMS_COLUMNS["qty"])

    merged = pd.merge(
        wcs_agg, wms_agg,
        on="ID",
        how="outer",        # MUST be outer — never lose IDs from either side
        indicator=True,
    )

    # Rename for the final report
    merged.rename(
        columns={WCS_COLUMNS["qty"]: "WCS Qty", WMS_COLUMNS["qty"]: "WMS Qty"},
        inplace=True,
    )

    # Fill NaN quantities with 0 for records only in one source
    merged["WCS Qty"] = merged["WCS Qty"].fillna(0).astype(int)
    merged["WMS Qty"] = merged["WMS Qty"].fillna(0).astype(int)

    # Vectorized status assignment (no iterrows!)
    conditions = [
        merged["_merge"] == "left_only",
        merged["_merge"] == "right_only",
        merged["WCS Qty"] == merged["WMS Qty"],
    ]
    choices = ["Missing in WMS", "Missing in WCS", "Matched"]
    merged["Status"] = np.select(conditions, choices, default="Unmatched")

    # Detailed logging of results
    only_wcs = (merged["_merge"] == "left_only").sum()
    only_wms = (merged["_merge"] == "right_only").sum()
    qty_mismatch = (
        (merged["_merge"] == "both") & (merged["Status"] == "Unmatched")
    ).sum()
    matched = (merged["Status"] == "Matched").sum()

    if only_wcs:
        logger.warning("%d ID(s) found ONLY in WCS → Missing in WMS.", only_wcs)
    if only_wms:
        logger.warning("%d ID(s) found ONLY in WMS → Missing in WCS.", only_wms)
    if qty_mismatch:
        logger.warning(
            "%d ID(s) present in both but have QUANTITY MISMATCH.", qty_mismatch
        )

    logger.info(
        "Match summary — Total: %d | Matched: %d | Unmatched: %d | "
        "Missing in WMS: %d | Missing in WCS: %d",
        len(merged), matched, qty_mismatch, only_wcs, only_wms,
    )

    # Drop merge indicator before output
    merged.drop(columns=["_merge"], inplace=True)

    # Ensure correct column order
    return merged[["ID", "WCS Qty", "WMS Qty", "Status"]]


# ===========================  REPORT GENERATION  ===========================

def _compute_col_widths(df: pd.DataFrame) -> list:
    """
    Fast vectorized column-width calculation from a DataFrame.
    Avoids reading individual openpyxl cells (which is O(rows*cols) and very slow).
    """
    widths = []
    for col in df.columns:
        header_len = len(str(col))
        if len(df) > 0:
            data_len = int(df[col].astype(str).str.len().max())
        else:
            data_len = 0
        widths.append(max(header_len, data_len))
    return widths


def _format_sheet(ws, col_widths=None) -> None:
    """
    Apply professional formatting to any worksheet.
    col_widths: pre-computed list of ints (one per column) from _compute_col_widths.
    Passing widths avoids the O(rows*cols) cell-scan that freezes on large files.
    """
    # Header row only — fast (just N columns, not N*rows)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Column widths — use pre-computed pandas values, not cell-by-cell openpyxl reads
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_widths and col_idx <= len(col_widths):
            width = col_widths[col_idx - 1] + 4
        else:
            hdr = ws.cell(row=1, column=col_idx).value
            width = (len(str(hdr)) if hdr else 8) + 4
        ws.column_dimensions[col_letter].width = min(width, 60)

    # Autofilter + freeze top row
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _apply_conditional_formatting(ws, status_col_letter: str, max_row: int) -> None:
    """
    Apply Excel conditional formatting rules on the Status column.
    Uses real CF rules that scale with the file (not static cell fills).
    """
    cell_range = f"{status_col_letter}2:{status_col_letter}{max_row}"

    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal", formula=['"Matched"'], fill=GREEN_FILL
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal", formula=['"Unmatched"'], fill=RED_FILL
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal", formula=['"Missing in WMS"'], fill=ORANGE_FILL
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal", formula=['"Missing in WCS"'], fill=ORANGE_FILL
        ),
    )


def _build_summary(report_df: pd.DataFrame) -> pd.DataFrame:
    """Build summary statistics DataFrame for the Summary sheet."""
    total = len(report_df)
    matched = (report_df["Status"] == "Matched").sum()
    unmatched = (report_df["Status"] == "Unmatched").sum()
    missing_wms = (report_df["Status"] == "Missing in WMS").sum()
    missing_wcs = (report_df["Status"] == "Missing in WCS").sum()
    match_pct = round((matched / total * 100), 2) if total > 0 else 0.0
    total_wcs_qty = report_df["WCS Qty"].sum()
    total_wms_qty = report_df["WMS Qty"].sum()

    summary_data = {
        "Metric": [
            "Total Records",
            "Matched",
            "Unmatched",
            "Missing in WMS",
            "Missing in WCS",
            "Match %",
            "Total WCS Qty",
            "Total WMS Qty",
        ],
        "Value": [
            total,
            matched,
            unmatched,
            missing_wms,
            missing_wcs,
            f"{match_pct}%",
            total_wcs_qty,
            total_wms_qty,
        ],
    }
    return pd.DataFrame(summary_data)


def _build_missing_records(report_df: pd.DataFrame) -> pd.DataFrame:
    """Extract records that exist in only one system for the Missing_Records sheet."""
    mask = report_df["Status"].isin(["Missing in WMS", "Missing in WCS"])
    return report_df[mask][["ID", "WCS Qty", "WMS Qty", "Status"]].reset_index(drop=True)


def _build_audit_log(
    wcs_path: str, wms_path: str, report_df: pd.DataFrame,
    wcs_rows: int, wms_rows: int, dup_df: pd.DataFrame
) -> pd.DataFrame:
    """Build an audit log DataFrame capturing execution details."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    entries = [
        ("Timestamp", now),
        ("WCS File", wcs_path),
        ("WMS File", wms_path),
        ("WCS Rows Loaded", wcs_rows),
        ("WMS Rows Loaded", wms_rows),
        ("Duplicate IDs Found", len(dup_df)),
        ("Total Compared", len(report_df)),
        ("Matched", (report_df["Status"] == "Matched").sum()),
        ("Unmatched", (report_df["Status"] == "Unmatched").sum()),
        ("Missing in WMS", (report_df["Status"] == "Missing in WMS").sum()),
        ("Missing in WCS", (report_df["Status"] == "Missing in WCS").sum()),
        ("Execution Status", "SUCCESS"),
    ]
    return pd.DataFrame(entries, columns=["Item", "Detail"])


def generate_report(
    report_df: pd.DataFrame,
    output_path: str,
    dup_df: pd.DataFrame,
    wcs_path: str,
    wms_path: str,
    wcs_rows: int,
    wms_rows: int,
) -> None:
    """
    Generate a multi-sheet Excel workbook:
      1. Comparison_Report — main comparison with conditional formatting
      2. Summary — dashboard metrics
      3. Duplicate_IDs — audit of duplicates
      4. Missing_Records — records in only one system
      5. Audit_Log — execution metadata
    """
    output_path = str(output_path)

    # Build supplementary DataFrames
    summary_df = _build_summary(report_df)
    missing_df = _build_missing_records(report_df)
    audit_df = _build_audit_log(wcs_path, wms_path, report_df, wcs_rows, wms_rows, dup_df)

    # Write all sheets
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        report_df.to_excel(writer, index=False, sheet_name="Comparison_Report")
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        if not dup_df.empty:
            dup_df.to_excel(writer, index=False, sheet_name="Duplicate_IDs")
        else:
            pd.DataFrame(columns=["Source", "Duplicate ID", "Count"]).to_excel(
                writer, index=False, sheet_name="Duplicate_IDs"
            )
        if not missing_df.empty:
            missing_df.to_excel(writer, index=False, sheet_name="Missing_Records")
        else:
            pd.DataFrame(columns=["ID", "WCS Qty", "WMS Qty", "Status"]).to_excel(
                writer, index=False, sheet_name="Missing_Records"
            )
        audit_df.to_excel(writer, index=False, sheet_name="Audit_Log")

    # Pre-compute column widths using pandas (vectorized — no openpyxl cell scanning)
    _empty_dup = pd.DataFrame(columns=["Source", "Duplicate ID", "Count"])
    _empty_miss = pd.DataFrame(columns=["ID", "WCS Qty", "WMS Qty", "Status"])
    sheet_widths = {
        "Comparison_Report": _compute_col_widths(report_df),
        "Summary":           _compute_col_widths(summary_df),
        "Duplicate_IDs":     _compute_col_widths(dup_df if not dup_df.empty else _empty_dup),
        "Missing_Records":   _compute_col_widths(missing_df if not missing_df.empty else _empty_miss),
        "Audit_Log":         _compute_col_widths(audit_df),
    }

    # Re-open with openpyxl for formatting
    wb = load_workbook(output_path)

    # Format each sheet — pass pre-computed widths to avoid slow cell scans
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _format_sheet(ws, col_widths=sheet_widths.get(sheet_name))

    # Apply conditional formatting on Comparison_Report Status column
    ws_main = wb["Comparison_Report"]
    # Find status column letter
    status_col_letter = None
    for col_idx in range(1, ws_main.max_column + 1):
        if ws_main.cell(row=1, column=col_idx).value == "Status":
            status_col_letter = get_column_letter(col_idx)
            break
    if status_col_letter and ws_main.max_row > 1:
        _apply_conditional_formatting(ws_main, status_col_letter, ws_main.max_row)

    wb.save(output_path)
    logger.info("Report saved → %s", output_path)


# ===========================  EXECUTION REPORT  ============================

def print_execution_report(
    report_df: pd.DataFrame, wcs_rows: int, wms_rows: int, dup_count: int
) -> None:
    """Print a clean execution summary to console for warehouse teams."""
    matched = (report_df["Status"] == "Matched").sum()
    unmatched = (report_df["Status"] == "Unmatched").sum()
    missing_wms = (report_df["Status"] == "Missing in WMS").sum()
    missing_wcs = (report_df["Status"] == "Missing in WCS").sum()

    print("\n" + "=" * 50)
    print("  RECONCILIATION EXECUTION REPORT")
    print("=" * 50)
    print(f"  Loaded WCS:        {wcs_rows:,} rows")
    print(f"  Loaded WMS:        {wms_rows:,} rows")
    print(f"  Duplicate IDs:     {dup_count:,}")
    print(f"  Matched:           {matched:,}")
    print(f"  Unmatched:         {unmatched:,}")
    print(f"  Missing in WMS:    {missing_wms:,}")
    print(f"  Missing in WCS:    {missing_wcs:,}")
    print("-" * 50)
    print("  Final report generated successfully.")
    print("=" * 50 + "\n")


# ===========================  ORCHESTRATOR  ================================

def reconcile(
    wcs_path: str,
    wms_path: str,
    output_path: str = "Comparison_Report.xlsx",
    wms_header_row: Optional[int] = None,
    wcs_header_row: Optional[int] = None,
) -> None:
    """
    End-to-end reconciliation pipeline.

    Parameters
    ----------
    wcs_path : str
        Path to the WCS Excel file (e.g. PTL Close Container_22_May_2026.xlsx).
    wms_path : str
        Path to the WMS Excel file (e.g. Export__8_.xlsx).
    output_path : str
        Destination path for the generated report (default: Comparison_Report.xlsx).
    wms_header_row : int, optional
        0-indexed header row for the WMS file. Auto-detected if omitted.
    wcs_header_row : int, optional
        0-indexed header row for the WCS file. Auto-detected if omitted.
    """
    logger.info("=" * 60)
    logger.info("RECONCILIATION START")
    logger.info("=" * 60)

    # 1. Load (dtype=str + keep_default_na=False for barcode integrity)
    wcs_raw = load_excel(wcs_path, WCS_REQUIRED_COLS, header_row=wcs_header_row)
    wms_raw = load_excel(wms_path, WMS_REQUIRED_COLS, header_row=wms_header_row)

    wcs_rows = len(wcs_raw)
    wms_rows = len(wms_raw)

    # 2. Clean (advanced normalization layer)
    wcs_clean = clean_wcs(wcs_raw)
    wms_clean = clean_wms(wms_raw)

    # 3. Concatenate IDs
    wcs_with_id = add_wcs_id(wcs_clean)
    wms_with_id = add_wms_id(wms_clean)

    # 4. Validate keys
    validate_no_null_keys(wcs_with_id, "ID", "WCS")
    validate_no_null_keys(wms_with_id, "ID", "WMS")
    validate_id_format(wcs_with_id, "ID", "WCS")
    validate_id_format(wms_with_id, "ID", "WMS")
    validate_duplicates(wcs_with_id, "ID", "WCS")
    validate_duplicates(wms_with_id, "ID", "WMS")
    validate_row_counts(wcs_with_id, wms_with_id)

    # 4b. Build duplicate audit data
    dup_wcs = get_duplicate_details(wcs_with_id, "ID", "WCS")
    dup_wms = get_duplicate_details(wms_with_id, "ID", "WMS")
    dup_df = pd.concat([dup_wcs, dup_wms], ignore_index=True)

    # 5. Match (full outer join — never lose IDs from either side)
    report_df = match_datasets(wcs_with_id, wms_with_id)

    # 6. Generate multi-sheet report
    generate_report(
        report_df, output_path, dup_df,
        wcs_path, wms_path, wcs_rows, wms_rows,
    )

    # 7. Print execution summary
    print_execution_report(report_df, wcs_rows, wms_rows, len(dup_df))

    logger.info("=" * 60)
    logger.info("RECONCILIATION COMPLETE")
    logger.info("=" * 60)


# ===========================  CLI ENTRY  ===================================

def main() -> None:
    """
    Command-line interface.

    Usage:
        python reconcile.py <wcs_file> <wms_file> [output_file]

    Examples:
        python reconcile.py "PTL Close Container_22_May_2026.xlsx" "Export__8_.xlsx"
        python reconcile.py wcs.xlsx wms.xlsx my_report.xlsx
    """
    if len(sys.argv) < 3:
        print(
            "Usage: python reconcile.py <wcs_file> <wms_file> [output_file]\n"
            "\n"
            "Arguments:\n"
            "  wcs_file    Path to the WCS Excel file\n"
            "  wms_file    Path to the WMS Excel file\n"
            "  output_file (optional) Output report path (default: Comparison_Report.xlsx)\n"
        )
        sys.exit(1)

    wcs_file = sys.argv[1]
    wms_file = sys.argv[2]
    output_file = sys.argv[3] if len(sys.argv) > 3 else "Comparison_Report.xlsx"

    try:
        reconcile(wcs_file, wms_file, output_file)
    except Exception:
        logger.exception("Reconciliation failed")
        sys.exit(1)


if __name__ == "__main__":
    main()
